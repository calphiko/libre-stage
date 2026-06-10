"""
Gig management router.

Handles CRUD operations for gigs, set and setlist management, PDF
export of printed setlists and the gig schedule calculation.

Requires authentication. Create/update/delete operations additionally
require the ``editor`` or ``admin`` role.

Prefix: ``/gigs``  |  Tag: ``gigs``
"""

# libre-stage - Band rehearsal and gig management software
# Copyright (C) 2026  libre-stage contributors
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from fastapi import APIRouter, Depends, HTTPException, Response, Query
from fastapi.responses import StreamingResponse
from datetime import date, datetime, time, timedelta
import hashlib
import json
import logging
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from pathlib import Path

from typing import List, Literal

from backend.pdf.generator import SetlistPDF
from backend.services.setlist import SetlistService
from backend.utils.check_permissions import check_admin, check_editor
from backend.utils.pdf_palette import find_logo_path, resolve_schedule_palette
from backend.utils.setlist_timing import calculate_setlist_timing, serialize_timing_for_api

from backend import models, schemas, auth, app_config

import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas

import os
from dotenv import load_dotenv

router = APIRouter(
    prefix="/gigs", tags=["gigs"], dependencies=[Depends(auth.get_current_user_dep)]
)

logger = logging.getLogger("uvicorn.error")
# suppress progress polls to reduce log clutter
block_endpoints = ["/gigs/log"]

load_dotenv(".env")

MM_CHANNEL = os.getenv("MM_CHANNEL_GIGS")

class LogFilter(logging.Filter):  # pragma: no cover
    def filter(self, record):
        if record.args and len(record.args) >= 3:
            if record.args[2] in block_endpoints:  # type: ignore
                return False
        return True


uvicorn_logger = logging.getLogger("uvicorn.access")
uvicorn_logger.addFilter(LogFilter())


DEFAULT_SCHEDULE_PALETTE = {
    "bg": colors.HexColor("#0B1220"),
    "card": colors.HexColor("#111827"),
    "primary": colors.HexColor("#60A5FA"),
    "text": colors.HexColor("#E2E8F0"),
    "muted": colors.HexColor("#94A3B8"),
    "header_bg": colors.HexColor("#1E293B"),
    "row_alt": colors.HexColor("#0F172A"),
    "line": colors.HexColor("#334155"),
    "fixed": colors.HexColor("#123129"),
}

GENRE_PALETTE = [
    '#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6',
    '#06b6d4', '#14b8a6', '#f97316', '#ec4899', '#84cc16',
    '#6366f1', '#22c55e', '#f43f5e', '#eab308', '#0ea5e9',
    '#a855f7', '#ef6c00', '#00acc1', '#7cb342', '#d81b60',
    '#3949ab', '#00897b', '#c0ca33', '#5e35b1', '#039be5'
]


def _build_rollover_datetimes(base_date: date, values: list[tuple[str, time]]) -> list[tuple[str, datetime]]:
    """Build a monotonic timeline and roll over to next day after midnight."""
    result: list[tuple[str, datetime]] = []
    day_offset = 0
    last_time: time | None = None

    for label, t in values:
        if last_time is not None and t < last_time:
            day_offset += 1
        dt = datetime.combine(base_date + timedelta(days=day_offset), t)
        result.append((label, dt))
        last_time = t

    return result


def _fixed_schedule_points(gig: models.Gig) -> list[tuple[str, datetime]]:
    """Build fixed schedule timestamps from gig base data (naive UTC)."""
    fixed_points: list[tuple[str, datetime]] = []
    if gig.datum is None:
        return fixed_points

    base_date = gig.datum if isinstance(gig.datum, date) else date.fromisoformat(str(gig.datum))

    ordered_times = [
        ("Einlass", gig.doors),
        ("Beginn", gig.begin),
        ("Ende", gig.end),
    ]
    present_times = [(label, t) for label, t in ordered_times if t is not None]
    return _build_rollover_datetimes(base_date, present_times)


def _build_schedule_response(gig: models.Gig) -> schemas.GigScheduleOut:
    fixed_items = [
        schemas.GigScheduleItemOut(
            id=None,
            gig_id=gig.id,
            item_datetime=item_dt,
            was=label,
            wer="",
            wo="",
            is_fixed=True,
        )
        for label, item_dt in _fixed_schedule_points(gig)
    ]
    dynamic_items = [
        schemas.GigScheduleItemOut.model_validate(item, from_attributes=True)
        for item in gig.schedule_items
    ]

    all_items = fixed_items + dynamic_items
    all_items.sort(key=lambda item: (item.item_datetime, 0 if item.is_fixed else 1))
    return schemas.GigScheduleOut(items=all_items)


def _calculate_setlist_version(payload: dict) -> str:
    version_input = {
        "id": payload.get("id"),
        "begin": payload.get("begin"),
        "end": payload.get("end"),
        "sets": payload.get("sets", []),
    }
    canonical = json.dumps(version_input, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:16]


def _build_setlist_payload(gig: models.Gig) -> dict:
    payload = gig.to_dict()
    payload["timing"] = serialize_timing_for_api(calculate_setlist_timing(gig))
    payload["setlist_version"] = _calculate_setlist_version(payload)
    return payload


def _raise_if_schedule_conflict(
    db: Session,
    gig: models.Gig,
    item_datetime: datetime,
    exclude_item_id: int | None = None,
) -> None:
    existing_query = db.query(models.GigScheduleItem).filter(
        models.GigScheduleItem.gig_id == gig.id,
        models.GigScheduleItem.item_datetime == item_datetime,
    )
    if exclude_item_id is not None:
        existing_query = existing_query.filter(models.GigScheduleItem.id != exclude_item_id)

    if existing_query.first() is not None:
        raise HTTPException(
            status_code=409,
            detail="Zeitpunkt kollidiert mit einem vorhandenen Ablaufplan-Eintrag.",
        )

    for label, fixed_dt in _fixed_schedule_points(gig):
        if fixed_dt == item_datetime:
            raise HTTPException(
                status_code=409,
                detail=f"Zeitpunkt kollidiert mit festem Eintrag: {label}.",
            )


def _build_schedule_pdf(gig: models.Gig) -> bytes:
    """Render schedule items (fixed + dynamic) to a printable portrait PDF."""
    schedule = _build_schedule_response(gig)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=0)
    page_width, page_height = A4

    left_margin = 36
    right_margin = page_width - 36
    bottom_margin = 52
    table_header_height = 20
    row_padding = 4
    row_line_height = 12

    table_width = right_margin - left_margin
    time_col_width = 96
    was_col_width = 180
    wer_col_width = 110
    wo_col_width = max(90, table_width - time_col_width - was_col_width - wer_col_width)
    was_col_width = max(140, table_width - time_col_width - wer_col_width - wo_col_width)

    columns = [
        ("Zeit", left_margin, time_col_width),
        ("Was", left_margin + time_col_width, was_col_width),
        ("Wer", left_margin + time_col_width + was_col_width, wer_col_width),
        (
            "Wo",
            left_margin + time_col_width + was_col_width + wer_col_width,
            wo_col_width,
        ),
    ]
    footer_prefix = "Generiert mit libreStage | "
    footer_domain = "pakleds-patentoffice.de"
    footer_url = "https://pakleds-patentoffice.de"
    footer_text = f"{footer_prefix}{footer_domain}"

    root_dir = Path(__file__).resolve().parents[2]
    logo_path = find_logo_path({"root_dir": root_dir})
    logo_reader: ImageReader | None = None
    logo_size: tuple[float, float] | None = None

    palette = resolve_schedule_palette(
        {
            "default_palette": DEFAULT_SCHEDULE_PALETTE,
            "logo_path": logo_path,
        }
    )

    def _get_logo_reader() -> ImageReader | None:
        nonlocal logo_reader, logo_size
        if logo_reader is not None:
            return logo_reader

        if logo_path is None:
            return None

        try:
            logo_reader = ImageReader(str(logo_path))
            logo_size = logo_reader.getSize()
            return logo_reader
        except Exception:  # pragma: no cover - logo rendering should not break export
            logger.warning("Could not load logo for schedule PDF", exc_info=True)
            return None

    def _format_time(value: time | None) -> str:
        return value.strftime("%H:%M") if value else "-"

    def _wrap_text(text: str, max_width: float, font_name: str, font_size: int) -> list[str]:
        content = (text or "-").strip() or "-"
        words = content.split()
        if not words:
            return ["-"]

        lines: list[str] = []
        current = ""
        for word in words:
            trial = f"{current} {word}" if current else word
            if pdfmetrics.stringWidth(trial, font_name, font_size) <= max_width:
                current = trial
                continue

            if current:
                lines.append(current)
                current = ""

            # Break very long tokens that do not fit into one column line.
            if pdfmetrics.stringWidth(word, font_name, font_size) <= max_width:
                current = word
                continue

            chunk = ""
            for char in word:
                trial_chunk = f"{chunk}{char}"
                if pdfmetrics.stringWidth(trial_chunk, font_name, font_size) <= max_width:
                    chunk = trial_chunk
                else:
                    if chunk:
                        lines.append(chunk)
                    chunk = char
            current = chunk

        if current:
            lines.append(current)
        return lines

    def draw_table_header(y: float) -> float:
        pdf.setFillColor(palette["header_bg"])
        pdf.roundRect(left_margin, y - table_header_height + 4, right_margin - left_margin, table_header_height, 4, stroke=0, fill=1)
        pdf.setFont("Helvetica-Bold", 10)
        pdf.setFillColor(palette["text"])
        for title, x, _ in columns:
            pdf.drawString(x + 4, y - 10, title)
        pdf.setStrokeColor(palette["line"])
        pdf.setLineWidth(0.8)
        pdf.line(left_margin, y - table_header_height, right_margin, y - table_header_height)
        return y - table_header_height - 6

    def draw_watermark() -> None:
        img = _get_logo_reader()
        if img is None or not logo_size:
            return

        img_width, img_height = logo_size
        max_width = page_width * 0.55
        max_height = page_height * 0.55
        ratio = min(max_width / img_width, max_height / img_height)
        render_w = img_width * ratio
        render_h = img_height * ratio
        x = (page_width - render_w) / 2
        y = (page_height - render_h) / 2 - 28

        pdf.saveState()
        # Keep watermark subtle so table data stays readable.
        if hasattr(pdf, "setFillAlpha"):
            pdf.setFillAlpha(0.05)
            pdf.setStrokeAlpha(0.05)
        pdf.drawImage(
            img,
            x,
            y,
            width=render_w,
            height=render_h,
            preserveAspectRatio=True,
            mask="auto",
        )
        pdf.restoreState()

    def draw_footer() -> None:
        pdf.setFillColor(palette["muted"])
        pdf.setFont("Helvetica", 8)
        footer_y = 20
        footer_font = "Helvetica"
        footer_size = 8
        full_width = pdfmetrics.stringWidth(footer_text, footer_font, footer_size)
        prefix_width = pdfmetrics.stringWidth(footer_prefix, footer_font, footer_size)
        start_x = (page_width - full_width) / 2

        pdf.drawString(start_x, footer_y, footer_prefix)
        pdf.setFillColor(palette["primary"])
        pdf.drawString(start_x + prefix_width, footer_y, footer_domain)
        pdf.linkURL(
            footer_url,
            (start_x + prefix_width, footer_y - 2, start_x + full_width, footer_y + footer_size + 1),
            relative=0,
            thickness=0,
        )

    def draw_header() -> float:
        pdf.setFillColor(palette["bg"])
        pdf.rect(0, 0, page_width, page_height, stroke=0, fill=1)

        draw_watermark()
        draw_footer()

        card_height = 134
        card_y = page_height - 26 - card_height
        pdf.setFillColor(palette["card"])
        pdf.roundRect(left_margin, card_y, right_margin - left_margin, card_height, 8, stroke=0, fill=1)
        pdf.setStrokeColor(palette["line"])
        pdf.setLineWidth(0.8)
        pdf.roundRect(left_margin, card_y, right_margin - left_margin, card_height, 8, stroke=1, fill=0)

        pdf.setFillColor(palette["primary"])
        pdf.rect(left_margin, card_y + card_height - 8, right_margin - left_margin, 8, stroke=0, fill=1)

        img = _get_logo_reader()
        if img is not None and logo_size:
            try:
                img_width, img_height = logo_size
                max_logo_width = 150
                max_logo_height = 50
                ratio = min(max_logo_width / img_width, max_logo_height / img_height)
                render_w = img_width * ratio
                render_h = img_height * ratio
                pdf.drawImage(
                    img,
                    right_margin - 14 - render_w,
                    card_y + card_height - 20 - render_h,
                    width=render_w,
                    height=render_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:  # pragma: no cover - logo rendering should not break export
                logger.warning("Could not render logo in schedule PDF", exc_info=True)

        title_x = left_margin + 14
        title_y = card_y + card_height - 26
        pdf.setFillColor(palette["text"])
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(title_x, title_y, "Ablaufplan")
        pdf.setFont("Helvetica", 16)
        pdf.setFillColor(palette["muted"])
        pdf.drawString(title_x, title_y - 20, f"Gig: {gig.name}")
        pdf.setFont("Helvetica", 11)
        date_str = gig.datum.strftime("%d.%m.%Y") if gig.datum else "-"
        base_fields = [
            f"Datum: {date_str}",
            f"Art: {gig.kind_of_gig or '-'}",
            f"Veranstalter: {gig.organizer or '-'}",
            f"Ort: {gig.venue or '-'}",
            f"Einlass/Beginn/Ende: {_format_time(gig.doors)} / {_format_time(gig.begin)} / {_format_time(gig.end)}",
            f"Export: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ]
        info_y = title_y - 36
        pdf.setFillColor(palette["text"])
        pdf.setFont("Helvetica", 11)
        for field in base_fields:
            pdf.drawString(title_x, info_y, field)
            info_y -= 13

        return draw_table_header(card_y - 11)

    schedule_font_name = "Helvetica"
    schedule_font_size = 9
    font_ascent = pdfmetrics.getAscent(schedule_font_name)
    font_descent = pdfmetrics.getDescent(schedule_font_name)
    baseline_center_offset = ((font_ascent + font_descent) / 2000.0) * schedule_font_size

    y = draw_header()
    pdf.setFont(schedule_font_name, schedule_font_size)
    row_index = 0

    for item in schedule.items:
        dt_str = item.item_datetime.strftime("%d.%m.%Y %H:%M")
        was_str = item.was + (" [fix]" if item.is_fixed else "")

        wrapped = {
            "Zeit": _wrap_text(dt_str, columns[0][2] - 4, "Helvetica", 9),
            "Was": _wrap_text(was_str, columns[1][2] - 4, "Helvetica", 9),
            "Wer": _wrap_text(item.wer or "-", columns[2][2] - 4, "Helvetica", 9),
            "Wo": _wrap_text(item.wo or "-", columns[3][2] - 4, "Helvetica", 9),
        }
        row_lines = max(len(lines) for lines in wrapped.values())
        row_height = row_lines * row_line_height + row_padding * 2

        if y - row_height < bottom_margin:
            pdf.showPage()
            y = draw_header()
            pdf.setFont(schedule_font_name, schedule_font_size)

        row_top = y + 4
        row_bottom = row_top - row_height
        if item.is_fixed:
            fill = palette["fixed"]
        elif row_index % 2:
            fill = palette["row_alt"]
        else:
            fill = None

        if fill is not None:
            pdf.setFillColor(fill)
            pdf.rect(left_margin, row_bottom, right_margin - left_margin, row_height, stroke=0, fill=1)

        for idx in range(row_lines):
            line_center = row_top - row_padding - (idx * row_line_height) - (row_line_height / 2)
            line_y = line_center - baseline_center_offset
            for title, x, _ in columns:
                lines = wrapped[title]
                if idx < len(lines):
                    if title == "Zeit":
                        pdf.setFillColor(palette["muted"])
                    else:
                        pdf.setFillColor(palette["text"])
                    pdf.drawString(x + 4, line_y, lines[idx])

        y -= row_height
        pdf.setStrokeColor(palette["line"])
        pdf.setLineWidth(0.5)
        pdf.line(left_margin, y + 2, right_margin, y + 2)
        row_index += 1

    pdf.save()
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/", response_model=List[schemas.GigOut])
def list_gigs(
        db: Session = Depends(auth.get_db),
        current_user=Depends(auth.get_current_user),
        jahr: int = Query(None, description="Das Jahr der Gigs"),
):
    logger.info("Fetching gigs from database")
    query = db.query(models.Gig).order_by(models.Gig.datum.desc())
    if jahr is not None:
        # Wir gehen davon aus, dass 'datum' als 'YYYY-MM-DD' (Text) gespeichert ist
        query = query.filter(models.Gig.datum.startswith(str(jahr)))
    return query.all()


@router.get("/statistics", response_model=schemas.SeasonStatistics)
def get_season_statistics(
    jahr: int = Query(None, description="Jahr für die Saisonstatistik"),
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user),
):
    """Gibt aggregierte Statistiken für alle Gigs eines Jahres zurück."""
    from datetime import date

    query = db.query(models.Gig).order_by(models.Gig.datum.asc())
    if jahr is not None:
        query = query.filter(models.Gig.datum.startswith(str(jahr)))
    gigs = query.all()

    gig_count = len(gigs)
    played_gig_count = sum(1 for gig in gigs if gig.datum and gig.datum < date.today())
    total_songs = 0
    unique_song_ids = set()
    skipped_count = 0
    inserted_count = 0
    feedback_values = []
    song_play_counts = {}  # song_id -> count
    gigs_overview = []
    genre_counts = {}  # genre -> count
    genre_timeline = []

    for gig in gigs:
        gig_songs = 0
        gig_skipped = 0
        gig_inserted = 0
        gig_feedbacks = []
        gig_genre_counts = {}

        for gigset in gig.sets:
            set_obj = gigset.set
            if not set_obj:
                continue
            for setsong in set_obj.songs:
                total_songs += 1
                gig_songs += 1
                if setsong.id_song:
                    unique_song_ids.add(setsong.id_song)
                    song_play_counts[setsong.id_song] = song_play_counts.get(setsong.id_song, 0) + 1
                if setsong.uebersprungen:
                    skipped_count += 1
                    gig_skipped += 1
                if setsong.eingeschoben:
                    inserted_count += 1
                    gig_inserted += 1
                if setsong.feedback is not None:
                    feedback_values.append(setsong.feedback)
                    gig_feedbacks.append(setsong.feedback)
                # Genre zählen
                if setsong.song and setsong.song.genre:
                    genre = setsong.song.genre.strip()
                    if genre:
                        genre_counts[genre] = genre_counts.get(genre, 0) + 1
                        gig_genre_counts[genre] = gig_genre_counts.get(genre, 0) + 1

        if gig_genre_counts:
            genre_timeline.append(schemas.GenreTimelinePoint(
                label=gig.name,
                date=gig.datum.strftime('%Y-%m-%d') if gig.datum else None,
                kind_of_gig=gig.kind_of_gig,
                genre_counts=gig_genre_counts,
                total=sum(gig_genre_counts.values()),
            ))

        gig_feedback_avg = round(sum(gig_feedbacks) / len(gig_feedbacks), 2) if gig_feedbacks else None
        gigs_overview.append(schemas.GigOverviewEntry(
            gig_id=gig.id,
            gig_name=gig.name,
            gig_date=gig.datum.strftime('%Y-%m-%d') if gig.datum else '',
            song_count=gig_songs,
            skipped_count=gig_skipped,
            inserted_count=gig_inserted,
            feedback_avg=gig_feedback_avg,
        ))

    feedback_distribution = {}
    for fv in feedback_values:
        feedback_distribution[fv] = feedback_distribution.get(fv, 0) + 1
    feedback_avg = round(sum(feedback_values) / len(feedback_values), 2) if feedback_values else None

    # Top 10 Songs
    top_song_entries = sorted(song_play_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    top_songs = []
    for song_id, count in top_song_entries:
        song = db.query(models.Song).get(song_id)
        if song:
            top_songs.append(schemas.TopSongEntry(
                song_id=song.id,
                title=song.title,
                interpret=song.interpret or '',
                count=count,
            ))

    return schemas.SeasonStatistics(
        jahr=jahr,
        gig_count=gig_count,
        played_gig_count=played_gig_count,
        total_songs=total_songs,
        unique_songs=len(unique_song_ids),
        skipped_count=skipped_count,
        inserted_count=inserted_count,
        feedback_count=len(feedback_values),
        feedback_avg=feedback_avg,
        feedback_distribution=feedback_distribution,
        genre_distribution=genre_counts,
        genre_timeline=genre_timeline,
        top_songs=top_songs,
        gigs_overview=gigs_overview,
    )


@router.get("/genre_palette", response_model=schemas.GenrePaletteOut)
def get_genre_palette(
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user),
):
    """Return a deterministic global genre->color map for all available genres."""
    genres_by_key: dict[str, str] = {}

    def _register_genre(value: str | None) -> None:
        if not isinstance(value, str):
            return
        text = value.strip()
        if not text:
            return
        key = text.casefold()
        genres_by_key.setdefault(key, text)

    # Prefer global configured genres when available.
    soft_config = app_config.get_soft_config()
    for entry in soft_config.get("genres", []):
        if isinstance(entry, str):
            _register_genre(entry)
        elif isinstance(entry, dict):
            _register_genre(entry.get("label") or entry.get("key"))

    # Add any legacy/ad-hoc genres present in DB songs.
    db_genres = db.query(models.Song.genre).filter(models.Song.genre.isnot(None)).distinct().all()
    for (genre,) in db_genres:
        _register_genre(genre)

    sorted_keys = sorted(genres_by_key.keys(), key=lambda k: k.casefold())
    palette = {
        genres_by_key[key]: GENRE_PALETTE[idx % len(GENRE_PALETTE)]
        for idx, key in enumerate(sorted_keys)
    }
    return schemas.GenrePaletteOut(palette=palette)


@router.post("/livemode_available_batch")
def get_livemode_available_batch(
    gig_ids: List[int],
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user)
):
    """
    Batch-Endpoint: Gibt Live-Mode-Status für mehrere Gigs auf einmal zurück.
    Body: { "gig_ids": [1, 2, 3, ...] }
    Returns: { "1": {...}, "2": {...}, ... }
    """
    from datetime import date

    # Normale User haben KEINEN Zugriff
    is_admin = check_admin(current_user)
    if not is_admin:
        # Gebe für alle Gigs "not available" zurück
        return {
            str(gig_id): {
                "available": False,
                "reason": "insufficient_permissions",
                "can_force": False
            }
            for gig_id in gig_ids
        }

    # Hole alle Gigs in einem Query
    gigs = db.query(models.Gig).filter(models.Gig.id.in_(gig_ids)).all()
    gigs_by_id = {gig.id: gig for gig in gigs}

    today = date.today()
    result = {}

    for gig_id in gig_ids:
        gig = gigs_by_id.get(gig_id)

        if not gig:
            result[str(gig_id)] = {
                "available": False,
                "reason": "gig_not_found",
                "can_force": False
            }
            continue

        # gig.datum ist bereits ein date-Objekt
        gig_date = gig.datum if isinstance(gig.datum, date) else date.fromisoformat(str(gig.datum))
        is_gig_day = today == gig_date

        result[str(gig_id)] = {
            "available": is_gig_day,
            "reason": "gig_day" if is_gig_day else "not_gig_day",
            "can_force": True,
            "gig_date": gig_date.isoformat()
        }

    return result

@router.get("/{gig_id}/livemode_available")
def is_livemode_available(
    gig_id: int,
    force: bool = Query(False, description="Editor/Admin override"),
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user)
):
    from datetime import date

    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    # Normale User haben KEINEN Zugriff
    is_editor_or_admin = check_editor(current_user)
    if not is_editor_or_admin:
        return {
            "available": False,
            "reason": "insufficient_permissions",
            "can_force": False
        }

    # Editor/Admin mit Force-Flag
    if force:
        return {
            "available": True,
            "forced": True,
            "reason": "manually_unlocked"
        }

    # Automatische Verfügbarkeit: Gig-Tag
    today = date.today()
    # gig.datum ist bereits ein date-Objekt
    gig_date = gig.datum if isinstance(gig.datum, date) else date.fromisoformat(str(gig.datum))
    is_gig_day = today == gig_date

    return {
        "available": is_gig_day,
        "reason": "gig_day" if is_gig_day else "not_gig_day",
        "can_force": True,
        "gig_date": gig_date.isoformat()
    }

@router.get("/{gig_id}/statistics", response_model=schemas.GigStatistics)
def get_gig_statistics(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user),
):
    """Gibt detaillierte Statistiken für einen einzelnen Gig zurück."""
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig nicht gefunden")

    song_count = 0
    skipped_count = 0
    inserted_count = 0
    feedback_values = []
    set_entries = []
    genre_counts = {}  # genre -> count
    genre_timeline = []

    for gigset in gig.sets:
        set_obj = gigset.set
        if not set_obj:
            continue

        set_songs = []
        set_feedbacks = []
        set_genre_counts = {}

        for setsong in sorted(set_obj.songs, key=lambda s: s.position):
            song = setsong.song
            title = song.title if song else "⚠️ Song gelöscht"
            interpret = (song.interpret or '') if song else ''
            song_id = setsong.id_song or 0

            song_count += 1
            if setsong.uebersprungen:
                skipped_count += 1
            if setsong.eingeschoben:
                inserted_count += 1
            if setsong.feedback is not None:
                feedback_values.append(setsong.feedback)
                set_feedbacks.append(setsong.feedback)
            # Genre zählen
            if song and song.genre:
                genre = song.genre.strip()
                if genre:
                    genre_counts[genre] = genre_counts.get(genre, 0) + 1
                    set_genre_counts[genre] = set_genre_counts.get(genre, 0) + 1

            set_songs.append(schemas.GigStatsSongEntry(
                song_id=song_id,
                title=title,
                interpret=interpret,
                position=setsong.position,
                feedback=setsong.feedback,
                uebersprungen=setsong.uebersprungen,
                eingeschoben=setsong.eingeschoben,
            ))

        set_feedback_avg = round(sum(set_feedbacks) / len(set_feedbacks), 2) if set_feedbacks else None
        if set_genre_counts:
            genre_timeline.append(schemas.GenreTimelinePoint(
                label=set_obj.setlist_name or set_obj.name or f"Set {len(set_entries) + 1}",
                date=gig.datum.strftime('%Y-%m-%d') if gig.datum else None,
                kind_of_gig=gig.kind_of_gig,
                genre_counts=set_genre_counts,
                total=sum(set_genre_counts.values()),
            ))
        set_entries.append(schemas.GigStatsSetEntry(
            set_name=set_obj.setlist_name or set_obj.name or f"Set {len(set_entries) + 1}",
            feedback_avg=set_feedback_avg,
            songs=set_songs,
        ))

    feedback_distribution = {}
    for fv in feedback_values:
        feedback_distribution[fv] = feedback_distribution.get(fv, 0) + 1
    feedback_avg = round(sum(feedback_values) / len(feedback_values), 2) if feedback_values else None

    return schemas.GigStatistics(
        gig_id=gig.id,
        gig_name=gig.name,
        gig_date=gig.datum.strftime('%Y-%m-%d') if gig.datum else '',
        song_count=song_count,
        skipped_count=skipped_count,
        inserted_count=inserted_count,
        feedback_count=len(feedback_values),
        feedback_avg=feedback_avg,
        feedback_distribution=feedback_distribution,
        genre_distribution=genre_counts,
        genre_timeline=genre_timeline,
        sets=set_entries,
    )


@router.get("/{gig_id}/schedule/", response_model=schemas.GigScheduleOut)
def get_gig_schedule(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user),
):
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")
    return _build_schedule_response(gig)


@router.get("/{gig_id}/schedule.pdf", response_class=Response)
def get_gig_schedule_pdf(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current_user=Depends(auth.get_current_user),
):
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    pdf_bytes = _build_schedule_pdf(gig)
    safe_name = "".join(ch if ch.isalnum() else "_" for ch in (gig.name or "gig"))
    headers = {"Content-Disposition": f"attachment; filename=Ablaufplan_{safe_name}_{gig_id}.pdf"}
    return Response(pdf_bytes, media_type="application/pdf", headers=headers)


@router.post("/{gig_id}/schedule/", response_model=schemas.GigScheduleOut)
def create_gig_schedule_item(
    gig_id: int,
    payload: schemas.GigScheduleItemIn,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user),
):
    if not check_editor(current):
        raise HTTPException(status_code=403, detail="Not enough permissions")

    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    _raise_if_schedule_conflict(db, gig, payload.item_datetime)

    item = models.GigScheduleItem(gig_id=gig_id, **payload.model_dump())
    db.add(item)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Zeitpunkt ist bereits belegt.")

    db.refresh(gig)
    return _build_schedule_response(gig)


@router.put("/{gig_id}/schedule/", response_model=schemas.GigScheduleOut)
def update_gig_schedule_bulk(
    gig_id: int,
    payload: schemas.GigScheduleBulkUpdateIn,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user),
):
    if not check_editor(current):
        raise HTTPException(status_code=403, detail="Not enough permissions")

    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    fixed_datetimes = {fixed_dt for _, fixed_dt in _fixed_schedule_points(gig)}
    payload_datetimes: set[datetime] = set()
    payload_ids: set[int] = set()

    for item in payload.items:
        if item.item_datetime in payload_datetimes:
            raise HTTPException(status_code=409, detail="Zeitpunkt doppelt im Ablaufplan-Payload.")
        payload_datetimes.add(item.item_datetime)

        if item.item_datetime in fixed_datetimes:
            raise HTTPException(status_code=409, detail="Zeitpunkt kollidiert mit einem festen Eintrag.")

        if item.id is not None:
            if item.id in payload_ids:
                raise HTTPException(status_code=409, detail="Eintrag-ID doppelt im Ablaufplan-Payload.")
            payload_ids.add(item.id)

    existing_items = db.query(models.GigScheduleItem).filter_by(gig_id=gig_id).all()
    existing_by_id = {item.id: item for item in existing_items}

    unknown_ids = [item_id for item_id in payload_ids if item_id not in existing_by_id]
    if unknown_ids:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    try:
        # Replace-all behavior: remove dynamic items not present in payload.
        for item in existing_items:
            if item.id not in payload_ids:
                db.delete(item)

        db.flush()

        updates = []
        for item_data in payload.items:
            if item_data.id is not None:
                updates.append((existing_by_id[item_data.id], item_data))

        # Move updated rows to guaranteed-free temporary datetimes first so swaps are possible.
        if updates:
            used_datetimes = {item.item_datetime for item in existing_items}
            used_datetimes.update(fixed_datetimes)
            used_datetimes.update(payload_datetimes)
            temp_anchor = datetime(9999, 12, 31, 23, 59, 59)
            temp_offset_seconds = 0

            for item, _ in updates:
                temp_dt = temp_anchor - timedelta(seconds=temp_offset_seconds)
                while temp_dt in used_datetimes:
                    temp_offset_seconds += 1
                    temp_dt = temp_anchor - timedelta(seconds=temp_offset_seconds)
                temp_offset_seconds += 1
                used_datetimes.add(temp_dt)
                item.item_datetime = temp_dt

            db.flush()

        for item_data in payload.items:
            if item_data.id is not None:
                item = existing_by_id[item_data.id]
                item.item_datetime = item_data.item_datetime
                item.was = item_data.was
                item.wer = item_data.wer
                item.wo = item_data.wo
            else:
                db.add(models.GigScheduleItem(gig_id=gig_id, **item_data.model_dump(exclude={"id"})))

        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Zeitpunkt ist bereits belegt.")

    db.refresh(gig)
    return _build_schedule_response(gig)


@router.put("/{gig_id}/schedule/{item_id}", response_model=schemas.GigScheduleOut)
def update_gig_schedule_item(
    gig_id: int,
    item_id: int,
    payload: schemas.GigScheduleItemIn,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user),
):
    if not check_editor(current):
        raise HTTPException(status_code=403, detail="Not enough permissions")

    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    item = db.query(models.GigScheduleItem).filter_by(id=item_id, gig_id=gig_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    _raise_if_schedule_conflict(db, gig, payload.item_datetime, exclude_item_id=item_id)

    for key, value in payload.model_dump().items():
        setattr(item, key, value)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Zeitpunkt ist bereits belegt.")

    db.refresh(gig)
    return _build_schedule_response(gig)


@router.delete("/{gig_id}/schedule/{item_id}", response_model=schemas.GigScheduleOut)
def delete_gig_schedule_item(
    gig_id: int,
    item_id: int,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user),
):
    if not check_editor(current):
        raise HTTPException(status_code=403, detail="Not enough permissions")

    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    item = db.query(models.GigScheduleItem).filter_by(id=item_id, gig_id=gig_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    db.delete(item)
    db.commit()
    db.refresh(gig)
    return _build_schedule_response(gig)


@router.put("/{gig_id}", response_model=schemas.GigOut)
def update_gig(
        gig_id: int,
        gig: schemas.GigIn,
        db: Session = Depends(auth.get_db),
        current=Depends(auth.get_current_user)
):
    if not check_editor(current):
        logger.error(f"Permission denied: User {current['user_name']} is not editor")
        raise HTTPException(status_code=401, detail="User role does not allow to update a gig!")

    logger.info(f"Updating gig in database with gig_id={gig_id}")
    db_gig = db.query(models.Gig).get(gig_id)

    if not db_gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    for k, v in gig.model_dump(exclude_unset=True).items():
        setattr(db_gig, k, v)
    db.commit()
    db.refresh(db_gig)

    return db_gig


def parse_name(full_name: str) -> dict:
    """
    Teilt einen vollständigen Namen in Vor- und Nachname auf.
    Falls nur ein Name vorhanden ist, wird dieser als Nachname verwendet.
    """
    if not full_name or not full_name.strip():
        return {'vorname': '', 'nachname': ''}

    # Mehrere Namen durch Komma/Semikolon getrennt -> nur ersten nehmen
    if ',' in full_name:
        full_name = full_name.split(',')[0].strip()
    elif ';' in full_name:
        full_name = full_name.split(';')[0].strip()
    elif '/' in full_name:
        full_name = full_name.split('/')[0].strip()

    parts = full_name.strip().split()

    if len(parts) == 0:
        return {'vorname': '', 'nachname': ''}
    elif len(parts) == 1:
        return {'vorname': '', 'nachname': parts[0]}
    else:
        # Letzter Teil ist Nachname, Rest ist Vorname
        return {
            'vorname': ' '.join(parts[:-1]),
            'nachname': parts[-1]
        }


@router.get("/{gig_id}/gemalist")
def download_gemalist(
        gig_id: int,
        db: Session = Depends(auth.get_db),
        current=Depends(auth.get_current_user)
):
    logger.info(f"Generating GEMA list for gig_id={gig_id}")

    # Gig laden mit allen Sets und Songs
    gig = db.query(models.Gig).filter(models.Gig.id == gig_id).first()
    if not gig:
        raise HTTPException(status_code=404, detail="Gig not found")

    # Excel-Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GEMA-Meldung"

    # Header-Informationen (Zeilen 1-18 wie in Vorlage)
    ws['A1'] = 'Excel-Vorlage zum Hochladen von Titeln'
    ws['C1'] = 'Version: 1.2'
    ws['K1'] = 'Live'
    ws['L1'] = 'Ja'
    ws['M1'] = 'Ja'
    ws['N1'] = 'F'

    # Gig-Informationen in Kopfbereich einfügen
    ws['A18'] = 'Setlist'

    # Header-Zeile (Zeile 19)
    headers = [
        'WERKNUMMER / WERKFASSUNGSNUMMER',
        'TITEL*',
        'SATZANGABE / SONSTIGE(R) Titel',
        'ANZAHL MUSIKER / SÄNGER*',
        'SPIELDAUER* (MM:SS)',
        'INTERPRET / KOMPONIST* (Nachname)',
        'INTERPRET / KOMPONIST (Vorname)',
        'TEXTDICHTER (Nachname)',
        'TEXTDICHTER (Vorname)',
        'BEARBEITER (Nachname)',
        'BEARBEITER (Vorname)',
        'VERLAG',
        'LIVE/TONTRÄGER',
        'VERÖFFENTLICHTES WERK',
        'POTPOURRI/FRAGMENT'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=19, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Datenzeilen ab Zeile 20
    current_row = 20

    # Alle Sets durchgehen (sortiert nach Position)
    for gigset in sorted(gig.sets, key=lambda x: x.position):
        set_obj = gigset.set

        # Alle Songs im Set durchgehen (sortiert nach Position)
        for setsong in sorted(set_obj.songs, key=lambda ss: ss.position):
            if setsong.uebersprungen:
                continue  # Übersprungene Songs nicht melden
            song = setsong.song

            # Spalte A: Werknummer (optional, leer lassen)
            ws.cell(row=current_row, column=1, value='')

            # Spalte B: TITEL* (Pflichtfeld)
            ws.cell(row=current_row, column=2, value=song.title or '')

            # Spalte C: SATZANGABE (optional)
            ws.cell(row=current_row, column=3, value='')

            # Spalte D: ANZAHL MUSIKER (optional, kann leer bleiben bei Pop/Rock)
            ws.cell(row=current_row, column=4, value='')

            # Spalte E: SPIELDAUER* (MM:SS) (Pflichtfeld)
            if song.duration:
                if isinstance(song.duration, time):
                    duration_str = song.duration.strftime("%M:%S")
                else:
                    duration_str = "00:00"
            else:
                duration_str = "00:00"
            ws.cell(row=current_row, column=5, value=duration_str)

            # Spalte F: INTERPRET / KOMPONIST* Nachname (Pflichtfeld)
            # Composer aufteilen in Vor- und Nachname
            composer_parts = parse_name(song.composer)
            ws.cell(row=current_row, column=6, value=composer_parts['nachname'])

            # Spalte G: INTERPRET / KOMPONIST Vorname
            ws.cell(row=current_row, column=7, value=composer_parts['vorname'])

            # Spalte H: TEXTDICHTER Nachname
            texter_parts = parse_name(song.texter)
            ws.cell(row=current_row, column=8, value=texter_parts['nachname'])

            # Spalte I: TEXTDICHTER Vorname
            ws.cell(row=current_row, column=9, value=texter_parts['vorname'])

            # Spalte J: BEARBEITER Nachname
            arrangement_parts = parse_name(song.arrangement)
            ws.cell(row=current_row, column=10, value=arrangement_parts['nachname'])

            # Spalte K: BEARBEITER Vorname
            ws.cell(row=current_row, column=11, value=arrangement_parts['vorname'])

            # Spalte L: VERLAG
            ws.cell(row=current_row, column=12, value=song.publisher or '')

            # Spalte M: LIVE/TONTRÄGER
            ws.cell(row=current_row, column=13, value='Live')

            # Spalte N: VERÖFFENTLICHTES WERK (J/N)
            ws.cell(row=current_row, column=14, value='J')

            # Spalte O: POTPOURRI/FRAGMENT (J/N)
            ws.cell(row=current_row, column=15, value='N')

            current_row += 1

    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 18

    # Excel in Memory speichern
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Dateiname mit Gig-Informationen
    filename = f"GEMA_Meldung_{gig.name}_{gig.datum.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{gig_id}/setlist.pdf", response_class=Response)
def download_setlist(
    gig_id: int,
    design: Literal["dark", "print"] = Query(
        "dark",
        description="Design-Variante fuer die Setlisten-PDF (dark oder print).",
    ),
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user),
):
    logger.info(f"Generating setlist PDF for gig_id={gig_id}")
    service = SetlistService(db)  # <- sync Session
    gig = service.load_gig(gig_id)
    #gig.debug_dump()

    if not gig:
        logger.error(f"Gig with id={gig_id} not found for PDF generation")
        raise HTTPException(status_code=404, detail="Gig not found")

    schedule = service.calc_schedule(gig)

    service.dump_gig_struct(gig)

    # Dynamische Sänger-Farben: Farben werden aus einer Palette zugewiesen
    SINGER_COLOR_PALETTE = [
        "#29B619", "#227FFF", "#E644C3", "#FF8C00", "#8B5CF6",
        "#059669", "#DC2626", "#CA8A04", "#0891B2", "#6366F1",
    ]

    singers = sorted({
        (setsong.song.singer_lead or "").replace("+"," ").replace(",", " ").split(" ")[0]
        for gs in gig.sets
        for setsong in gs.set.songs
        if setsong.song.singer_lead
    })

    singer_colors = {
        singer: SINGER_COLOR_PALETTE[i % len(SINGER_COLOR_PALETTE)]
        for i, singer in enumerate(singers)
    }

    pdf_bytes = SetlistPDF(gig, schedule, singer_colors, style_mode=design).build().getvalue()
    mode_suffix = "_druckfreundlich" if design == "print" else ""
    headers = {"Content-Disposition": f"inline; filename=Setliste_{gig.name}{mode_suffix}.pdf"}
    return Response(
        pdf_bytes,
        media_type="application/pdf",
        headers=headers
    )


@router.get("/{gig_id}/setlist", response_model=schemas.GigSetlistOut)
def get_gig_setlist(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    logger.info(f"Fetching gig setlist from database with gig_id={gig_id}")
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        logger.error(f"Gig with id={gig_id} not found")
        raise HTTPException(status_code=404, detail="Gig not found")

    # Prüfe auf korrupte SetSongs (ohne Song-Referenz) und entferne sie
    for gigset in gig.sets:
        set_obj = gigset.set
        corrupted_setsongs = [ss for ss in set_obj.songs if not ss.song]
        if corrupted_setsongs:
            logger.warning(f"Found {len(corrupted_setsongs)} corrupted SetSongs in Set {set_obj.id}, removing them")
            for ss in corrupted_setsongs:
                db.delete(ss)
            db.commit()

    return _build_setlist_payload(gig)


@router.get("/{gig_id}/forscore-setlist", response_class=Response)
def export_forscore_setlist(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    import plistlib
    logger.info(f"Generating forScore setlist for gig_id={gig_id}")
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        logger.error(f"Gig with id={gig_id} not found")
        raise HTTPException(status_code=404, detail="Gig not found")

    setlist_name = (gig.name or "libreStage Setlist").strip() or "libreStage Setlist"
    setlist_items = []
    for gigset in sorted(gig.sets, key=lambda x: x.position):
        set_obj = gigset.set
        setsonglist = sorted(set_obj.songs, key=lambda ss: ss.position)
        for setsong in setsonglist:
            if setsong.song:
                setlist_items.append({
                    "title": setsong.song.title,
                    "setlist": setlist_name,
                })

    try:
        xml_data = plistlib.dumps(setlist_items, fmt=plistlib.FMT_XML)
    except Exception as e:
        logger.error(f"Failed to generate PLIST for forScore: {e}")
        raise HTTPException(status_code=500, detail="Fehler bei der PLIST-Generierung")

    formatted_date = gig.datum.strftime('%Y-%m-%d') if gig.datum else "gig"
    safe_name = "".join(c for c in gig.name if c.isalnum() or c in ("-", "_", " ")).strip().replace(" ", "_")
    filename = f"Setlist-{formatted_date}-{safe_name}.4ss"

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return Response(
        content=xml_data,
        media_type="application/x-forscore-setlist",
        headers=headers
    )

# @app.put("/append_song_to_set", response_model=schemas.GigSetlistOut)
# def append_song_to_set(
#     data: schemas.SongToSetIn,
#     db: Session = Depends(auth.get_db),
#     current=Depends(auth.get_current_user)
# ):
#     db_gig = db.query(models.Gig).filter_by(id=data.gigId).first()
#     print(f"Appending {data.sondId} to {data.setId} on position {data.position}!")
#     return db_gig


@router.put("/{gig_id}/update_setlist/", response_model=schemas.GigSetlistOut)
def update_gig_setlist(
    gig_id: int,
    gig: schemas.GetSetlistIn,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    logger.info(f"Updating gig setlist in database with gig_id={gig_id}")
    db_gig = db.query(models.Gig).filter_by(id=gig_id).first()
    if not db_gig:
        logger.error(f"Gig with id={gig_id} not found for setlist update")
        raise HTTPException(status_code=404, detail="Gig not found")

    if not check_editor(current):
        logger.error(f"Permission denied: User {current['user_name']} is not editor")
        raise HTTPException(status_code=403, detail="Not enough permissions")

    current_payload = _build_setlist_payload(db_gig)
    if gig.setlist_version and gig.setlist_version != current_payload["setlist_version"]:
        logger.warning(
            "Setlist conflict for gig_id=%s by user=%s (client=%s db=%s)",
            gig_id,
            current.get("user_name"),
            gig.setlist_version,
            current_payload["setlist_version"],
        )
        raise HTTPException(
            status_code=409,
            detail={
                "code": "SETLIST_CONFLICT",
                "message": "Setliste wurde zwischenzeitlich geaendert. Bitte erneut versuchen.",
                "current_setlist": current_payload,
            },
        )

    # Aktuelle Sets des Gigs
    old_gigsets = db.query(models.GigSet).filter_by(id_gig=gig_id).all()
    existing_sets = {gs.set.id: gs.set for gs in old_gigsets}
    old_gigset_ids = [gs.id for gs in old_gigsets]

    # Input-Set-IDs und -Songs extrahieren
    input_set_ids = set(sd.set_id for sd in gig.sets if getattr(sd, 'set_id', None))

    # 1. Entferne nur GigSet-Verknüpfungen, die im neuen Input fehlen
    for gigset in old_gigsets:
        if gigset.set.id not in input_set_ids:
            db.delete(gigset)
    db.commit()

    # 2. Sets löschen, die nicht mehr genutzt werden
    for set_id in existing_sets:
        if set_id not in input_set_ids:
            db.query(models.SetSong).filter_by(id_set=set_id).delete()
            db.query(models.Set).filter_by(id=set_id).delete()
    db.commit()

    # 3. Update/Erstelle Sets und Songs
    new_gigsets = []
    for set_pos, set_data in enumerate(gig.sets, start=1):
        # Existierendes Set, falls vorhanden
        if getattr(set_data, 'set_id', None) and set_data.set_id in existing_sets:
            set_obj = existing_sets[set_data.set_id]
            # Update Eigenschaften des Sets
            set_obj.name = set_data.set_name
            set_obj.pause = time.fromisoformat(set_data.pause) if set_data.pause else None
            set_obj.setlist_name = set_data.setlist_name
            db.flush()

            # Speichere Live-Mode-Daten der bestehenden SetSongs
            old_setsongs = db.query(models.SetSong).filter_by(id_set=set_obj.id).all()
            livemode_data = {}  # Key: song_id, Value: {eingeschoben, uebersprungen, feedback}
            for old_ss in old_setsongs:
                livemode_data[old_ss.id_song] = {
                    'eingeschoben': old_ss.eingeschoben,
                    'uebersprungen': old_ss.uebersprungen,
                    'feedback': old_ss.feedback
                }

            # Update Songs im Set: Entferne alle alten, füge neue hinzu
            db.query(models.SetSong).filter_by(id_set=set_obj.id).delete()
        else:
            # Neues Set anlegen
            set_obj = models.Set(
                name=set_data.set_name,
                pause=time.fromisoformat(set_data.pause) if set_data.pause else None,
                setlist_name=set_data.setlist_name,
            )
            db.add(set_obj)
            db.flush()
            livemode_data = {}  # Keine alten Daten bei neuem Set

        # Jetzt Songs der Reihe nach neu anlegen
        for song_pos, song_data in enumerate(set_data.songs, start=1):
            song = db.query(models.Song).filter_by(id=song_data.song_id).first()
            if not song:
                raise HTTPException(status_code=400, detail=f"Song not found")

            # Hole alte Live-Mode-Daten wenn vorhanden
            old_data = livemode_data.get(song.id, {})

            setsong = models.SetSong(
                id_set=set_obj.id,
                id_song=song.id,
                position=song_pos,
                eingeschoben=old_data.get('eingeschoben'),
                uebersprungen=old_data.get('uebersprungen'),
                feedback=old_data.get('feedback')
            )
            db.add(setsong)

        # GigSet-Verknüpfung erneuern
        gigset = db.query(models.GigSet).filter_by(id_gig=db_gig.id, id_set=set_obj.id).first()
        if not gigset:
            gigset = models.GigSet(
                id_gig=db_gig.id,
                id_set=set_obj.id,
                position=set_pos
            )
            db.add(gigset)
        else:
            gigset.position = set_pos
        new_gigsets.append(gigset)

    db.commit()
    db.refresh(db_gig)
    return _build_setlist_payload(db_gig)

@router.post("/")
def create_new_gig(
    gig: schemas.NewGig,
    jahr = Query(None, description="Das Jahr der Gigs"),
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    logger.info(f"Creating new gig in database with name={gig.name}")
    if not check_editor(current):
        logger.error(f"Permission denied: User {current['user_name']} is not admin or editor")
        raise HTTPException(status_code=401, detail="User role does not allow to create a new gig!")

    gig_to_add = models.Gig(
        name=gig.name,
        datum=gig.datum,
        kind_of_gig=gig.kind_of_gig,
        organizer=gig.organizer,
        venue=gig.venue,
        doors=gig.doors,
        begin=gig.begin,
        end=gig.end,
        status=gig.status,
        publish=False,
    )
    db.add(gig_to_add)
    db.commit()

    query = db.query(models.Gig).order_by(models.Gig.datum.desc())
    if jahr is not None:
        query = query.filter(models.Gig.datum.startswith(str(jahr)))

    # Mattermost notification
    try:
        from backend.utils import mattermost
        message = f":mega: Es wurde soeben ein neuer Gig erstellt:\n\tName: **{gig.name}**\n\tDatum: **{gig.datum.strftime('%d.%m.%Y')}**"
        if gig.begin:
            message += f"\n\tBeginn: **{gig.begin.strftime('%H:%M Uhr')}**"
        if gig.end:
            message += f"\n\tEnde: **{gig.end.strftime('%H:%M Uhr')}**"
        if gig.venue:
            message += f"\n\tOrt: **{gig.venue}**"
        if gig.organizer:
            message += f"\n\tVeranstalter: **{gig.organizer}**"
        mattermost.send_mm_message(channel=MM_CHANNEL, text=message)
    except Exception as e:
        logger.error(f"Failed to send Mattermost message: {e}")

    return query.all()

@router.delete("/{gig_id}", response_model=List[schemas.GigOut])
def delete_gig(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    logger.info(f"Deleting gig in database with gig_id={gig_id}")
    if not check_editor(current):
        logger.error(f"Permission denied: User {current['user_name']} is not admin or editor")
        raise HTTPException(status_code=401, detail="User role does not allow to delete a gig!")

    gig_to_del = db.query(models.Gig).get(gig_id)
    if not gig_to_del:
        logger.error(f"Gig with id={gig_id} not found for deletion")
        raise HTTPException(status_code=404, detail="Gig not found")

    for gs in list(gig_to_del.sets):
        logger.info("Deleting Gigset")
        # DELETE ALL SETS
        db.delete(gs)
    # DELETE GIG
    db.delete(gig_to_del)
    db.commit()

    query = db.query(models.Gig).order_by(models.Gig.datum.desc())
    jahr = None
    if jahr is not None:
        query = query.filter(models.Gig.datum.startswith(str(jahr))) # pragma: no cover

    # Mattermost notification
    try:
        from backend.utils import mattermost
        message = f":mega: Der Gig **{gig_to_del.name}** am **{gig_to_del.datum.strftime('%d.%m.%Y')}** wurde abgesagt."
        mattermost.send_mm_message(channel=MM_CHANNEL, text=message)
    except Exception as e:
        logger.error(f"Failed to send Mattermost message: {e}")

    return query.all()


@router.get("/{gig_id}/setlist_available")
def is_setlist_available(
    gig_id: int,
    db: Session = Depends(auth.get_db),
    current=Depends(auth.get_current_user)
):
    logger.info(f"Checking if setlist is available for gig_id={gig_id}")
    gig = db.query(models.Gig).get(gig_id)
    if not gig:
        logger.error(f"Gig with id={gig_id} not found for setlist availability check")
        raise HTTPException(status_code=404, detail="Gig not found")

    set_available = len(gig.sets) > 0
    setsong_available = any(len(gs.set.songs) > 0 for gs in gig.sets)
    is_available = set_available and setsong_available
    logger.info(f"Setlist availability for gig_id={gig_id}: {is_available}")
    return {"setlist_available": is_available}
